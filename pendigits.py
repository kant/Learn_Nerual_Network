# -*- coding: utf-8 -*-
"""
Created on Mon Nov  9 16:50:27 2020

@author: Xu Mingjun
"""

import torch
import torch.nn as nn
import numpy as np
from openpyxl import load_workbook
import torch.nn.functional as F

#继承nn.Module构建自己的简单神经网络SNet
class SNet(torch.nn.Module):
    #构建三层全连接网络
    def __init__(self, D_in, H1, H2, D_out):
        super(SNet, self).__init__()
        #定义每层的结构（每层是线性的）
        self.linear1 = torch.nn.Linear(D_in, H1)
        self.linear2 = torch.nn.Linear(H1, H2)
        self.linear3 = torch.nn.Linear(H2, D_out)
        
    #使用SNet会自动运行forward（前向传播）方法，方法连接各个隐藏层，并产生非线性
    def forward(self, x):
        y_1 = F.relu(self.linear1(x))
        y_2 = F.relu(self.linear2(y_1))
        y_pred = self.linear3(y_2);
        return y_pred
    
#网络向量初始化方法
def weights_init(m):
    #如果网络是线性层的话
    if isinstance(m, (nn.Linear)):
        nn.init.xavier_normal_(m.weight)    #初始‘w’，用Glorot初始化方法，可以用其他方法代替
        nn.init.constant_(m.bias, 0.0)  #把偏置‘b’初始化为常数0

#python 函数入口
if __name__ == "__main__":
    
    #读取表格中的数据
    workbook = load_workbook('dataset\pendigits.xlsx')
    booksheet = workbook.get_sheet_by_name('Sheet1')
    
    #初始化Ndarray用于存放表中变量
    data = np.zeros((10992,17))
    
    #读取sheet1中的数据到numpy的矩阵data中
    for i in range(10992):    
        for j in range(17): 
            data[i][j] = booksheet.cell(row=i+2, column=j+1).value
    
    #默认对Ndarray的第0维打乱（对样本的顺序进行打乱）
    np.random.shuffle(data)  
          
    #分出10000个训练数据
    train_label = data[0:10000,-1]      #[0:10000]不包含10000，-1是取最后一列
    train_feature = data[0:10000,0:-1]  #0：-1是从0取到最后第二列
    #分出992个测试数据
    test_label = data[10000:10992,-1]
    test_feature = data[10000:10992,0:-1]
    
    #对label进行onehot编码（多分类）
    train_label_oh = np.eye(10)[train_label.astype(int)]
    test_label_oh = np.eye(10)[test_label.astype(int)]
    
    #对feature归一化处理
    train_feature = train_feature / train_feature.max(axis=0)
    test_feature = test_feature / test_feature.max(axis=0)
    
    #训练和测试进行类型转换
    train_feature_t = torch.from_numpy(train_feature)   #由numpy转换而来是torch.float64
    train_label_t = torch.from_numpy(train_label_oh)
    
    train_feature_t_f = torch.tensor(train_feature_t,dtype=torch.float32)   #转换为torch.float32，因为要和之后模型预测的类型匹配
    train_label_t_f = torch.tensor(train_label_t,dtype=torch.float32)
    
    test_feature_t = torch.from_numpy(test_feature)
    test_label_t = torch.from_numpy(test_label_oh)
    
    test_feature_t_f = torch.tensor(test_feature_t,dtype=torch.float32)
    test_label_t_f = torch.tensor(test_label_t,dtype=torch.float32)
    
    #######################训练过程##############################
    
    #输入维度，隐藏层神经元个数，输出维度
    D_in, H1, H2, D_out = 16, 20, 15, 10
    
    #实例化一个用于预测的网络
    model = SNet(D_in, H1, H2, D_out)
    
    #定义损失函数，使用均方误差
    loss_fn = nn.MSELoss(reduction='sum')
    #设置学习率
    learning_rate = 1e-3
    #优化器
    optimizer = torch.optim.Adam(model.parameters(), lr=learning_rate)
    #网络参数初始化（上面定义的方法）
    model.apply(weights_init)
    
    #10000轮迭代，可更改
    for t in range(10000): 
        # 执行model.forward()前向传播
        y_pred = model(train_feature_t_f)
        
        #计算本次迭代的误差
        loss = loss_fn(y_pred, train_label_t_f)
        print("训练次数：",t,"\tloss =",loss.item())
        
        #清零梯度
        optimizer.zero_grad()
        #Backward pass反向传播
        loss.backward()
        #更新参数
        optimizer.step()
    
    #######################测试过程##############################
    
    model.eval()    #保证BN和dropout不发生变化
    
    cnt = 0 #初始化正确的计数值
    
    #输入训练集得到测试结果
    test_out = model(test_feature_t_f)
    _, test_out_np= torch.max(test_out,1)   #onehot解码，返回值第一个是最大值（不需要），第二个是最大值的序号

    #迭代922个测试样本输出和统计    
    for test_i in range(992):
    
        print("No.",test_i,"\npre:",test_out_np.numpy()[test_i],"\nGT:",test_label[test_i])
        print("****************")
        if test_out_np.numpy()[test_i] == test_label[test_i]:
            #print("correct")
            cnt += 1
    
    #正确率计算
    correct_rate = cnt/992.0
    print("correct_rate:",correct_rate)
